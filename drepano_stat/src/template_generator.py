"""Génération du classeur Excel standardisé de DrepanoStat."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = [
    "type_condition",
    "groupe",
    "dilution",
    "temoin",
    "repetition",
    "N",
    "D",
    "commentaire",
]
DEFAULT_DILUTIONS = ["Solution mère", "1_10", "1_100", "1_1000"]
TYPE_CONDITIONS = ["Extrait", "Témoin"]

HEADER_FILL = PatternFill("solid", fgColor="8B1E3F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
INVALID_FILL = PatternFill("solid", fgColor="F4CCCC")


def _clean_values(values: Iterable[str], field_name: str) -> list[str]:
    cleaned = [str(value).strip() for value in values if str(value).strip()]
    cleaned = list(dict.fromkeys(cleaned))
    if not cleaned:
        raise ValueError(f"{field_name} doit contenir au moins une valeur.")
    return cleaned


def _add_list_validation(worksheet, cell_range: str, formula: str) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "Choisissez une valeur dans la liste proposée."
    validation.errorTitle = "Valeur non valide"
    validation.prompt = "Sélectionnez une valeur dans le menu déroulant."
    validation.promptTitle = "Valeurs autorisées"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    worksheet.add_data_validation(validation)
    validation.add(cell_range)


def generate_excel_template(
    study_name: str,
    product_name: str,
    groups: Iterable[str],
    dilutions: Iterable[str],
    controls: Iterable[str],
    repetitions: int,
    generated_at: datetime | None = None,
) -> bytes:
    """Retourne en mémoire un modèle XLSX prêt à être téléchargé."""
    study_name = study_name.strip()
    product_name = product_name.strip()
    groups = _clean_values(groups, "groups")
    dilutions = _clean_values(dilutions, "dilutions")
    controls = _clean_values(controls, "controls")
    if not study_name:
        raise ValueError("study_name ne peut pas être vide.")
    if not isinstance(repetitions, int) or isinstance(repetitions, bool) or repetitions < 1:
        raise ValueError("repetitions doit être un entier supérieur ou égal à 1.")

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Données"
    parameters_sheet = workbook.create_sheet("Paramètres")
    instructions_sheet = workbook.create_sheet("Instructions")

    data_sheet.append(HEADERS)
    for group in groups:
        for dilution in dilutions:
            for repetition in range(1, repetitions + 1):
                data_sheet.append(["Extrait", group, dilution, "", repetition, "", "", ""])
    for control in controls:
        for repetition in range(1, repetitions + 1):
            data_sheet.append(["Témoin", "", "", control, repetition, "", "", ""])

    for cell in data_sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = f"A1:H{data_sheet.max_row}"
    widths = [18, 24, 18, 20, 12, 12, 12, 35]
    for index, width in enumerate(widths, start=1):
        data_sheet.column_dimensions[get_column_letter(index)].width = width
    for row in range(2, data_sheet.max_row + 1):
        for column in (6, 7, 8):
            data_sheet.cell(row, column).fill = INPUT_FILL

    generated_at = generated_at or datetime.now().astimezone()
    parameters_sheet.append(["Paramètre", "Valeur"])
    parameters_sheet.append(["Nom de l’étude", study_name])
    parameters_sheet.append(["Plante / produit", product_name])
    parameters_sheet.append(["Nombre de répétitions", repetitions])
    parameters_sheet.append(["Date de génération", generated_at.strftime("%Y-%m-%d %H:%M %Z")])
    parameters_sheet.append([])
    parameters_sheet.append(["Types de condition", "Groupes", "Dilutions", "Témoins"])
    list_start_row = 8
    max_values = max(len(TYPE_CONDITIONS), len(groups), len(dilutions), len(controls))
    for index in range(max_values):
        parameters_sheet.append(
            [
                TYPE_CONDITIONS[index] if index < len(TYPE_CONDITIONS) else "",
                groups[index] if index < len(groups) else "",
                dilutions[index] if index < len(dilutions) else "",
                controls[index] if index < len(controls) else "",
            ]
        )
    for row_number in (1, 7):
        for cell in parameters_sheet[row_number]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
    parameters_sheet.column_dimensions["A"].width = 24
    parameters_sheet.column_dimensions["B"].width = 35
    parameters_sheet.column_dimensions["C"].width = 22
    parameters_sheet.column_dimensions["D"].width = 24
    parameters_sheet.freeze_panes = "A8"

    parameter_sheet_name = quote_sheetname(parameters_sheet.title)
    named_lists = {
        "TypesCondition": ("A", len(TYPE_CONDITIONS)),
        "GroupesExtraits": ("B", len(groups)),
        "DilutionsEtude": ("C", len(dilutions)),
        "TemoinsEtude": ("D", len(controls)),
    }
    for name, (column, value_count) in named_lists.items():
        reference = (
            f"{parameter_sheet_name}!${column}${list_start_row}:"
            f"${column}${list_start_row + value_count - 1}"
        )
        workbook.defined_names.add(DefinedName(name, attr_text=reference))

    last_data_row = data_sheet.max_row
    _add_list_validation(
        data_sheet,
        f"A2:A{last_data_row}",
        "=TypesCondition",
    )
    for column, range_name in (
        ("B", "GroupesExtraits"),
        ("C", "DilutionsEtude"),
        ("D", "TemoinsEtude"),
    ):
        _add_list_validation(
            data_sheet,
            f"{column}2:{column}{last_data_row}",
            f"={range_name}",
        )

    whole_number = DataValidation(
        type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True
    )
    whole_number.error = "N et D doivent être des nombres entiers positifs ou nuls."
    whole_number.errorTitle = "Comptage non valide"
    whole_number.showErrorMessage = True
    data_sheet.add_data_validation(whole_number)
    whole_number.add(f"F2:G{last_data_row}")
    data_sheet.conditional_formatting.add(
        f"A2:H{last_data_row}",
        FormulaRule(formula=['AND($A2="Extrait",OR($B2="",$C2=""))'], fill=INVALID_FILL),
    )
    data_sheet.conditional_formatting.add(
        f"A2:H{last_data_row}",
        FormulaRule(formula=['AND($A2="Témoin",$D2="")'], fill=INVALID_FILL),
    )

    instructions = [
        "Instructions de remplissage",
        "1. Ne modifiez pas les noms des colonnes ni ceux des feuilles.",
        "2. Pour un Extrait, conservez type_condition, groupe, dilution et répétition.",
        "3. Pour un Témoin, conservez type_condition, témoin et répétition.",
        "4. Saisissez uniquement des nombres entiers positifs ou nuls dans N et D.",
        "5. N et D doivent être renseignés et leur somme doit être supérieure à zéro.",
        "6. La colonne commentaire est facultative.",
        "7. Une cellule rouge signale une combinaison incomplète.",
        "8. Enregistrez le fichier au format .xlsx avant de le réimporter dans DrepanoStat.",
    ]
    for instruction in instructions:
        instructions_sheet.append([instruction])
    instructions_sheet["A1"].fill = HEADER_FILL
    instructions_sheet["A1"].font = HEADER_FONT
    instructions_sheet.column_dimensions["A"].width = 105
    for row in instructions_sheet.iter_rows():
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    instructions_sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
